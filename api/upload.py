"""
Vercel Serverless Function: POST /api/upload
Accepts a Zerodha F&O P&L Excel file, parses it, computes all analytics, and returns JSON.
"""
import json
import re
import io
import os
import math
import uuid
import urllib.request
import urllib.error
from http.server import BaseHTTPRequestHandler
from collections import defaultdict

import openpyxl


# ─── Symbol Parsing Helpers ────────────────────────────────────────────────────

INDEX_UNDERLYINGS = {"NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX", "BANKEX"}


def get_underlying(symbol: str) -> str:
    match = re.match(r"^([A-Z&]+?)\d{2}", symbol)
    return match.group(1) if match else symbol


def get_instrument_type(symbol: str) -> str:
    if symbol.endswith("CE"):
        return "CE"
    elif symbol.endswith("PE"):
        return "PE"
    elif "FUT" in symbol:
        return "FUT"
    return "UNKNOWN"


def get_strike(symbol: str) -> str:
    match = re.search(r"(\d{3,6})(CE|PE)$", symbol)
    return match.group(1) if match else ""


def is_index(underlying: str) -> bool:
    return underlying.upper() in INDEX_UNDERLYINGS


# ─── Excel Parsing ─────────────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # --- F&O Sheet ---
    ws = wb["F&O"]

    # Find summary values by scanning for labels in column B
    summary = {}
    charges_detail = []
    header_row = None
    date_range_text = ""

    for row_num in range(1, min(ws.max_row + 1, 50)):
        cell_b = ws.cell(row=row_num, column=2).value
        cell_c = ws.cell(row=row_num, column=3).value

        if cell_b is None:
            continue

        cell_b_str = str(cell_b).strip()

        # Date range
        if "P&L Statement" in cell_b_str:
            date_range_text = cell_b_str

        # Summary values
        if cell_b_str == "Charges" and cell_c is not None and isinstance(cell_c, (int, float)):
            summary["charges"] = float(cell_c)
        elif cell_b_str == "Other Credit & Debit" and cell_c is not None:
            summary["other_credit_debit"] = float(cell_c)
        elif cell_b_str == "Realized P&L" and cell_c is not None:
            summary["realized_pnl"] = float(cell_c)
        elif cell_b_str == "Unrealized P&L" and cell_c is not None:
            summary["unrealized_pnl"] = float(cell_c)

        # Individual charge items (after "Account Head" header)
        if cell_b_str == "Account Head":
            # Next rows are charge items until empty
            for cr in range(row_num + 1, row_num + 20):
                cb = ws.cell(row=cr, column=2).value
                cc = ws.cell(row=cr, column=3).value
                if cb is None or str(cb).strip() == "":
                    break
                label = str(cb).strip()
                # Clean up label: remove " - Z" suffix
                label = re.sub(r"\s*-\s*Z$", "", label)
                amount = float(cc) if cc is not None else 0.0
                if amount > 0:
                    charges_detail.append({"name": label, "amount": amount})

        # Data header row
        if cell_b_str == "Symbol" and str(ws.cell(row=row_num, column=3).value or "").strip() == "ISIN":
            header_row = row_num

    if header_row is None:
        raise ValueError("Could not find data header row in F&O sheet")

    # Parse date range
    date_range = ""
    dr_match = re.search(r"from (\d{4}-\d{2}-\d{2}) to (\d{4}-\d{2}-\d{2})", date_range_text)
    if dr_match:
        from datetime import datetime
        d1 = datetime.strptime(dr_match.group(1), "%Y-%m-%d")
        d2 = datetime.strptime(dr_match.group(2), "%Y-%m-%d")
        date_range = f"{d1.strftime('%d %b %Y')} - {d2.strftime('%d %b %Y')}"

    # Parse position data
    positions = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        symbol = ws.cell(row=row_num, column=2).value
        if symbol is None or str(symbol).strip() == "":
            continue

        symbol = str(symbol).strip()
        qty = float(ws.cell(row=row_num, column=4).value or 0)
        buy_value = float(ws.cell(row=row_num, column=5).value or 0)
        sell_value = float(ws.cell(row=row_num, column=6).value or 0)
        realized_pnl = float(ws.cell(row=row_num, column=7).value or 0)
        realized_pnl_pct = float(ws.cell(row=row_num, column=8).value or 0)
        prev_close = float(ws.cell(row=row_num, column=9).value or 0)
        open_qty = float(ws.cell(row=row_num, column=10).value or 0)
        open_qty_type = str(ws.cell(row=row_num, column=11).value or "")
        open_value = float(ws.cell(row=row_num, column=12).value or 0)
        unrealized_pnl = float(ws.cell(row=row_num, column=13).value or 0)
        unrealized_pnl_pct = float(ws.cell(row=row_num, column=14).value or 0)

        underlying = get_underlying(symbol)
        inst_type = get_instrument_type(symbol)
        strike = get_strike(symbol)

        pos = {
            "symbol": symbol,
            "underlying": underlying,
            "instrument_type": inst_type,
            "strike": strike,
            "is_index": is_index(underlying),
            "quantity": qty,
            "buy_value": buy_value,
            "sell_value": sell_value,
            "realized_pnl": realized_pnl,
            "realized_pnl_pct": realized_pnl_pct,
            "prev_close": prev_close,
            "open_quantity": open_qty,
            "open_quantity_type": open_qty_type,
            "open_value": open_value,
            "unrealized_pnl": unrealized_pnl,
            "unrealized_pnl_pct": unrealized_pnl_pct,
        }
        positions.append(pos)

    # Split positions
    closed = [p for p in positions if p["quantity"] > 0]
    open_pos = [p for p in positions if p["quantity"] == 0 and p["open_quantity"] > 0]

    # --- Other Debits and Credits Sheet ---
    other_entries = []
    if "Other Debits and Credits" in wb.sheetnames:
        ws2 = wb["Other Debits and Credits"]
        odc_header_row = None
        for row_num in range(1, ws2.max_row + 1):
            cb = ws2.cell(row=row_num, column=2).value
            if cb and str(cb).strip() == "Particulars":
                odc_header_row = row_num
                break
        if odc_header_row:
            for row_num in range(odc_header_row + 1, ws2.max_row + 1):
                particulars = ws2.cell(row=row_num, column=2).value
                if particulars is None:
                    continue
                posting_date = str(ws2.cell(row=row_num, column=3).value or "")
                debit = float(ws2.cell(row=row_num, column=4).value or 0)
                credit = float(ws2.cell(row=row_num, column=5).value or 0)
                other_entries.append({
                    "particulars": str(particulars),
                    "posting_date": posting_date,
                    "debit": debit,
                    "credit": credit,
                })

    wb.close()

    return {
        "summary": summary,
        "charges_detail": charges_detail,
        "closed_positions": closed,
        "open_positions": open_pos,
        "other_entries": other_entries,
        "date_range": date_range,
        "total_positions": len(positions),
    }


# ─── Analytics Computation ──────────────────────────────────────────────────────

def compute_analytics(parsed: dict) -> dict:
    summary = parsed["summary"]
    charges_detail = parsed["charges_detail"]
    closed = parsed["closed_positions"]
    open_pos = parsed["open_positions"]
    other_entries = parsed["other_entries"]

    realized_pnl = summary.get("realized_pnl", 0)
    unrealized_pnl = summary.get("unrealized_pnl", 0)
    total_charges = summary.get("charges", 0)
    other_cd = summary.get("other_credit_debit", 0)
    net_after_charges = realized_pnl - total_charges + other_cd

    # ─── Overview ───────────────────────────────────────────────────────────
    winners = [p for p in closed if p["realized_pnl"] > 0]
    losers = [p for p in closed if p["realized_pnl"] < 0]
    breakeven = [p for p in closed if p["realized_pnl"] == 0]

    win_rate = (len(winners) / len(closed) * 100) if closed else 0

    total_buy = sum(p["buy_value"] for p in closed)
    total_sell = sum(p["sell_value"] for p in closed)

    avg_winner = (sum(p["realized_pnl"] for p in winners) / len(winners)) if winners else 0
    avg_loser = (sum(p["realized_pnl"] for p in losers) / len(losers)) if losers else 0
    win_loss_ratio = (abs(avg_winner) / abs(avg_loser)) if avg_loser != 0 else 0

    # Symbols by type
    type_counts = defaultdict(int)
    for p in closed:
        type_counts[p["instrument_type"]] += 1

    # P&L Distribution (histogram bins)
    pnl_values = [p["realized_pnl"] for p in closed]
    pnl_distribution = _make_histogram(pnl_values, num_bins=15, prefix="₹")

    # Return % Distribution
    return_pcts = [p["realized_pnl_pct"] for p in closed if p["buy_value"] > 0]
    return_distribution = _make_histogram(return_pcts, num_bins=15, suffix="%")

    # Symbol-wise P&L table (sorted by absolute P&L desc)
    symbol_pnl = sorted(closed, key=lambda p: abs(p["realized_pnl"]), reverse=True)
    symbol_pnl_table = [{
        "symbol": p["symbol"],
        "underlying": p["underlying"],
        "type": p["instrument_type"],
        "buy_value": p["buy_value"],
        "sell_value": p["sell_value"],
        "realized_pnl": p["realized_pnl"],
        "pnl_pct": p["realized_pnl_pct"],
    } for p in symbol_pnl]

    overview = {
        "net_realized_pnl": realized_pnl,
        "unrealized_pnl": unrealized_pnl,
        "total_charges": total_charges,
        "net_after_charges": net_after_charges,
        "win_rate": round(win_rate, 1),
        "symbols_traded": len(closed),
        "symbols_by_type": dict(type_counts),
        "total_buy_value": total_buy,
        "total_sell_value": total_sell,
        "winners": len(winners),
        "losers": len(losers),
        "breakeven": len(breakeven),
        "avg_winner": round(avg_winner, 0),
        "avg_loser": round(avg_loser, 0),
        "win_loss_ratio": round(win_loss_ratio, 2),
        "pnl_distribution": pnl_distribution,
        "return_pct_distribution": return_distribution,
        "symbol_wise_pnl": symbol_pnl_table,
    }

    # ─── Performance Attribution ────────────────────────────────────────────
    # Group by underlying
    underlying_pnl = defaultdict(lambda: {"pnl": 0, "buy_value": 0, "sell_value": 0, "count": 0, "wins": 0})
    for p in closed:
        u = p["underlying"]
        underlying_pnl[u]["pnl"] += p["realized_pnl"]
        underlying_pnl[u]["buy_value"] += p["buy_value"]
        underlying_pnl[u]["sell_value"] += p["sell_value"]
        underlying_pnl[u]["count"] += 1
        if p["realized_pnl"] > 0:
            underlying_pnl[u]["wins"] += 1

    sorted_underlyings = sorted(underlying_pnl.items(), key=lambda x: x[1]["pnl"], reverse=True)

    # Top 3 winners/losers
    top3_winners = sorted_underlyings[:3]
    top3_losers = sorted_underlyings[-3:]
    top3_winners_impact = sum(v["pnl"] for _, v in top3_winners)
    top3_losers_impact = sum(v["pnl"] for _, v in top3_losers)
    top3_winners_pct = (top3_winners_impact / realized_pnl * 100) if realized_pnl else 0
    top3_losers_pct = (abs(top3_losers_impact) / realized_pnl * 100) if realized_pnl else 0

    # Concentration risk (top 5)
    top5_abs = sorted(underlying_pnl.items(), key=lambda x: abs(x[1]["pnl"]), reverse=True)[:5]
    concentration = sum(abs(v["pnl"]) for _, v in top5_abs)
    total_abs_pnl = sum(abs(v["pnl"]) for v in underlying_pnl.values())
    concentration_pct = (concentration / total_abs_pnl * 100) if total_abs_pnl else 0

    # Top contributors and detractors
    top_contributors = [{"underlying": u, "pnl": v["pnl"]} for u, v in sorted_underlyings[:5]]
    top_detractors = [{"underlying": u, "pnl": v["pnl"]} for u, v in sorted_underlyings[-5:]]

    # Waterfall (top 20 by absolute pnl)
    waterfall_sorted = sorted(underlying_pnl.items(), key=lambda x: abs(x[1]["pnl"]), reverse=True)[:20]
    # Sort: positives first (desc), then negatives (asc)
    waterfall_pos = sorted([x for x in waterfall_sorted if x[1]["pnl"] >= 0], key=lambda x: x[1]["pnl"], reverse=True)
    waterfall_neg = sorted([x for x in waterfall_sorted if x[1]["pnl"] < 0], key=lambda x: x[1]["pnl"])
    waterfall_ordered = waterfall_pos + waterfall_neg
    cumulative = 0
    waterfall = []
    for u, v in waterfall_ordered:
        cumulative += v["pnl"]
        waterfall.append({"underlying": u, "pnl": round(v["pnl"], 0), "cumulative": round(cumulative, 0)})

    # Pareto analysis (winning underlyings only)
    winning_underlyings = [(u, v) for u, v in sorted_underlyings if v["pnl"] > 0]
    total_winning = sum(v["pnl"] for _, v in winning_underlyings)
    cum_pct = 0
    pareto = []
    pareto_80_index = 0
    for i, (u, v) in enumerate(winning_underlyings):
        cum_pct += (v["pnl"] / total_winning * 100) if total_winning else 0
        pareto.append({"underlying": u, "pnl": round(v["pnl"], 0), "cumulative_pct": round(cum_pct, 1)})
        if cum_pct >= 80 and pareto_80_index == 0:
            pareto_80_index = i

    # Win/Loss asymmetry by underlying
    winning_u = [v["pnl"] for _, v in sorted_underlyings if v["pnl"] > 0]
    losing_u = [v["pnl"] for _, v in sorted_underlyings if v["pnl"] < 0]
    avg_win_u = (sum(winning_u) / len(winning_u)) if winning_u else 0
    avg_loss_u = (sum(losing_u) / len(losing_u)) if losing_u else 0
    asymmetry = (abs(avg_win_u) / abs(avg_loss_u)) if avg_loss_u != 0 else 0

    # Contribution table
    contribution_table = []
    for u, v in sorted_underlyings:
        contribution_pct = (v["pnl"] / realized_pnl * 100) if realized_pnl else 0
        contribution_table.append({
            "underlying": u,
            "num_symbols": v["count"],
            "pnl": round(v["pnl"], 0),
            "contribution_pct": round(contribution_pct, 1),
            "buy_value": round(v["buy_value"], 0),
            "sell_value": round(v["sell_value"], 0),
        })

    performance = {
        "top_winners_impact": round(top3_winners_impact, 0),
        "top_winners_pct": round(top3_winners_pct, 1),
        "top_winners_names": ", ".join(u for u, _ in top3_winners),
        "top_losers_impact": round(abs(top3_losers_impact), 0),
        "top_losers_pct": round(top3_losers_pct, 1),
        "top_losers_names": ", ".join(u for u, _ in top3_losers),
        "concentration_risk": round(concentration, 0),
        "concentration_pct": round(concentration_pct, 1),
        "concentration_detail": ", ".join(u for u, _ in top5_abs),
        "top_contributors": top_contributors,
        "top_detractors": top_detractors,
        "waterfall": waterfall,
        "pareto": pareto,
        "pareto_80_index": pareto_80_index,
        "avg_win": round(avg_win_u, 0),
        "avg_loss": round(avg_loss_u, 0),
        "win_loss_asymmetry": round(asymmetry, 2),
        "contribution_table": contribution_table,
    }

    # ─── Instrument Breakdown ───────────────────────────────────────────────
    futures = [p for p in closed if p["instrument_type"] == "FUT"]
    options = [p for p in closed if p["instrument_type"] in ("CE", "PE")]
    calls = [p for p in closed if p["instrument_type"] == "CE"]
    puts = [p for p in closed if p["instrument_type"] == "PE"]
    index_fo = [p for p in closed if p["is_index"]]

    futures_pnl = sum(p["realized_pnl"] for p in futures)
    options_pnl = sum(p["realized_pnl"] for p in options)
    calls_pnl = sum(p["realized_pnl"] for p in calls)
    puts_pnl = sum(p["realized_pnl"] for p in puts)
    index_pnl = sum(p["realized_pnl"] for p in index_fo)

    # P&L by underlying table
    pnl_by_underlying = []
    for u, v in sorted_underlyings:
        win_rate_u = (v["wins"] / v["count"] * 100) if v["count"] else 0
        avg_return = (v["pnl"] / v["buy_value"] * 100) if v["buy_value"] else 0
        pnl_by_underlying.append({
            "underlying": u,
            "num_symbols": v["count"],
            "pnl": round(v["pnl"], 0),
            "win_rate": round(win_rate_u, 0),
            "avg_return": round(avg_return, 1),
            "buy_value": round(v["buy_value"], 0),
        })

    # Capital vs returns scatter
    capital_returns = [{
        "underlying": u,
        "capital": round(v["buy_value"], 0),
        "return_pct": round((v["pnl"] / v["buy_value"] * 100) if v["buy_value"] else 0, 1),
    } for u, v in sorted_underlyings]

    instruments = {
        "futures_pnl": round(futures_pnl, 0),
        "options_pnl": round(options_pnl, 0),
        "calls_pnl": round(calls_pnl, 0),
        "puts_pnl": round(puts_pnl, 0),
        "index_fo_pnl": round(index_pnl, 0),
        "futures_count": len(futures),
        "options_count": len(options),
        "calls_count": len(calls),
        "puts_count": len(puts),
        "index_count": len(index_fo),
        "futures_vs_options": {"futures": round(futures_pnl, 0), "options": round(options_pnl, 0)},
        "calls_vs_puts": {"calls": round(calls_pnl, 0), "puts": round(puts_pnl, 0)},
        "index_vs_stock": {
            "index": round(index_pnl, 0),
            "stock": round(sum(p["realized_pnl"] for p in closed if not p["is_index"]), 0),
        },
        "pnl_by_underlying": pnl_by_underlying,
        "capital_vs_returns": capital_returns,
    }

    # ─── Charges & Costs ────────────────────────────────────────────────────
    charges_pct_pnl = (total_charges / abs(realized_pnl) * 100) if realized_pnl else 0
    total_turnover = total_buy + total_sell
    charges_pct_turnover = (total_charges / total_turnover * 100) if total_turnover else 0

    # Symbols to cover charges: count winning positions (sorted by pnl asc - smallest first) until cumulative >= charges
    sorted_winners = sorted(winners, key=lambda p: p["realized_pnl"])
    cum = 0
    symbols_to_cover = 0
    for p in sorted_winners:
        cum += p["realized_pnl"]
        symbols_to_cover += 1
        if cum >= total_charges:
            break

    # Charges breakdown with percentages
    total_charges_sum = sum(c["amount"] for c in charges_detail) if charges_detail else total_charges
    charges_breakdown = [{
        "name": c["name"],
        "amount": round(c["amount"], 0),
        "pct": round(c["amount"] / total_charges_sum * 100, 1) if total_charges_sum else 0,
    } for c in charges_detail]

    # Other debits and credits
    total_debits = sum(e["debit"] for e in other_entries)
    total_credits = sum(e["credit"] for e in other_entries)

    charges = {
        "total_charges": round(total_charges, 0),
        "charges_pct_pnl": round(charges_pct_pnl, 1),
        "charges_pct_turnover": round(charges_pct_turnover, 2),
        "symbols_to_cover": symbols_to_cover,
        "charges_breakdown": charges_breakdown,
        "gross_pnl": round(realized_pnl, 0),
        "net_pnl": round(net_after_charges, 0),
        "detailed_charges": charges_breakdown,
        "other_debits_credits": {
            "total_debits": round(total_debits, 0),
            "total_credits": round(total_credits, 0),
            "net": round(total_credits - total_debits, 0),
            "entries": other_entries,
        },
    }

    # ─── Open Portfolio ─────────────────────────────────────────────────────
    open_unrealized_pnl = sum(p["unrealized_pnl"] for p in open_pos)
    max_profit = max((p["unrealized_pnl"] for p in open_pos), default=0)
    max_loss = min((p["unrealized_pnl"] for p in open_pos), default=0)

    # Concentration by underlying
    open_by_underlying = defaultdict(float)
    for p in open_pos:
        open_by_underlying[p["underlying"]] += abs(p["open_value"])
    total_open_val = sum(open_by_underlying.values())
    concentration_list = sorted([{
        "underlying": u,
        "pct": round(v / total_open_val * 100, 1) if total_open_val else 0,
    } for u, v in open_by_underlying.items()], key=lambda x: x["pct"], reverse=True)

    # Unrealized by position
    unrealized_by_pos = sorted([{
        "symbol": p["symbol"],
        "underlying": p["underlying"],
        "pnl": round(p["unrealized_pnl"], 0),
    } for p in open_pos], key=lambda x: x["pnl"], reverse=True)

    # Open positions table
    open_table = [{
        "symbol": p["symbol"],
        "underlying": p["underlying"],
        "type": p["instrument_type"],
        "strike": p["strike"],
        "qty": int(p["open_quantity"]),
        "qty_type": p["open_quantity_type"],
        "buy_price": round(p["open_value"] / p["open_quantity"], 2) if p["open_quantity"] else 0,
        "open_value": round(p["open_value"], 0),
        "unrealized_pnl": round(p["unrealized_pnl"], 0),
        "pnl_pct": round(p["unrealized_pnl_pct"], 2),
        "prev_close": p["prev_close"],
    } for p in open_pos]

    open_portfolio = {
        "open_positions_count": len(open_pos),
        "total_unrealized_pnl": round(open_unrealized_pnl, 0),
        "total_open_value": round(total_open_val, 0),
        "max_unrealized_profit": round(max_profit, 0),
        "max_unrealized_loss": round(max_loss, 0),
        "concentration": concentration_list,
        "unrealized_by_position": unrealized_by_pos,
        "positions": open_table,
    }

    return {
        "metadata": {
            "date_range": parsed["date_range"],
            "total_symbols": parsed["total_positions"],
        },
        "overview": overview,
        "performance": performance,
        "instruments": instruments,
        "charges": charges,
        "open_portfolio": open_portfolio,
    }


def _make_histogram(values: list, num_bins: int = 15, prefix: str = "", suffix: str = "") -> list:
    if not values:
        return []
    min_val = min(values)
    max_val = max(values)
    if min_val == max_val:
        return [{"range": f"{prefix}{int(min_val)}{suffix}", "count": len(values)}]

    bin_width = (max_val - min_val) / num_bins
    # Round bin width to nice number
    magnitude = 10 ** math.floor(math.log10(abs(bin_width))) if bin_width != 0 else 1
    bin_width = math.ceil(bin_width / magnitude) * magnitude

    start = math.floor(min_val / bin_width) * bin_width
    bins = []
    current = start
    while current < max_val:
        bins.append(current)
        current += bin_width
    bins.append(current)

    counts = [0] * (len(bins) - 1)
    for v in values:
        for i in range(len(bins) - 1):
            if bins[i] <= v < bins[i + 1] or (i == len(bins) - 2 and v == bins[i + 1]):
                counts[i] += 1
                break

    result = []
    for i in range(len(counts)):
        lo = int(bins[i])
        hi = int(bins[i + 1])
        label = f"{prefix}{lo}{suffix} to {prefix}{hi}{suffix}"
        result.append({"range": label, "count": counts[i], "min": lo, "max": hi})

    return result


# ─── Supabase Storage (best-effort) ──────────────────────────────────────────────

def store_to_supabase(result, filename, file_bytes, request_headers):
    supabase_url = os.environ.get("SUPABASE_URL", "")
    supabase_key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not supabase_url or not supabase_key:
        return

    report_id = str(uuid.uuid4())
    storage_path = "{}_{}".format(report_id, filename)

    headers = {
        "apikey": supabase_key,
        "Authorization": "Bearer {}".format(supabase_key),
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    row = {
        "id": report_id,
        "filename": filename,
        "ip_address": request_headers.get("X-Forwarded-For", "unknown"),
        "user_agent": request_headers.get("User-Agent", "unknown"),
        "date_range": result.get("metadata", {}).get("date_range", ""),
        "total_symbols": result.get("metadata", {}).get("total_symbols", 0),
        "net_realized_pnl": result.get("overview", {}).get("net_realized_pnl", 0),
        "report_json": result,
        "file_storage_path": storage_path,
    }

    # Insert report row
    body = json.dumps(row).encode("utf-8")
    req = urllib.request.Request(
        "{}/rest/v1/reports".format(supabase_url),
        data=body,
        headers=headers,
        method="POST",
    )
    try:
        urllib.request.urlopen(req, timeout=10)
    except Exception:
        pass

    # Upload raw xlsx to Storage bucket
    storage_headers = {
        "apikey": supabase_key,
        "Authorization": "Bearer {}".format(supabase_key),
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    storage_req = urllib.request.Request(
        "{}/storage/v1/object/xlsx-uploads/{}".format(supabase_url, storage_path),
        data=file_bytes,
        headers=storage_headers,
        method="POST",
    )
    try:
        urllib.request.urlopen(storage_req, timeout=15)
    except Exception:
        pass


# ─── HTTP Handler ───────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        content_type = self.headers.get("Content-Type", "")

        if "multipart/form-data" not in content_type:
            self.send_response(400)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "Expected multipart/form-data"}).encode())
            return

        # Parse multipart form data
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length)

        # Extract boundary
        boundary = None
        for part in content_type.split(";"):
            part = part.strip()
            if part.startswith("boundary="):
                boundary = part[len("boundary="):].strip('"')
                break

        if not boundary:
            self.send_response(400)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "No boundary in multipart data"}).encode())
            return

        # Find file content between boundaries
        boundary_bytes = ("--" + boundary).encode()
        parts = body.split(boundary_bytes)

        file_bytes = None
        filename = "unknown.xlsx"
        for part in parts:
            if b"Content-Disposition" in part and b"filename=" in part:
                # Extract filename
                header_end = part.find(b"\r\n\r\n")
                if header_end == -1:
                    continue
                header_section = part[:header_end].decode("utf-8", errors="replace")
                fn_match = re.search(r'filename="([^"]+)"', header_section)
                if fn_match:
                    filename = fn_match.group(1)

                file_content = part[header_end + 4:]
                # Remove trailing \r\n-- if present
                if file_content.endswith(b"\r\n"):
                    file_content = file_content[:-2]
                if file_content.endswith(b"--\r\n"):
                    file_content = file_content[:-4]
                file_bytes = file_content
                break

        if file_bytes is None:
            self.send_response(400)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "No file found in upload"}).encode())
            return

        if not filename.endswith(".xlsx"):
            self.send_response(400)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": "Only .xlsx files are accepted"}).encode())
            return

        try:
            parsed = parse_excel(file_bytes)
            result = compute_analytics(parsed)
            result["metadata"]["filename"] = filename

            # Store to Supabase (best-effort)
            try:
                store_to_supabase(result, filename, file_bytes, {
                    "X-Forwarded-For": self.headers.get("X-Forwarded-For", ""),
                    "User-Agent": self.headers.get("User-Agent", ""),
                })
            except Exception:
                pass

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps(result).encode())
        except Exception as e:
            self.send_response(422)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"error": f"Failed to parse file: {str(e)}"}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
